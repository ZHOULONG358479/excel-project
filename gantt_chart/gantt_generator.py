from matplotlib import font_manager, rcParams
from datetime import datetime
import hashlib
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from pathlib import Path
import matplotlib.dates as mdates
import textwrap
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.ticker import FixedLocator
import numpy as np

# -----------------------------
# 0️⃣ 中文字体适配（macOS优先）
# -----------------------------
chinese_fonts = ["PingFang SC", "Songti SC", "STSong", "Heiti SC", "SimHei", "Arial Unicode MS"]
available_fonts = {f.name for f in font_manager.fontManager.ttflist}
for font in chinese_fonts:
    if font in available_fonts:
        rcParams["font.family"] = font
        break
else:
    raise RuntimeError("系统中未找到可用的中文字体")
rcParams["axes.unicode_minus"] = False

# -----------------------------
# 主题样式（工程汇报风格）
# -----------------------------
PROCESS_COLOR_PALETTE = [
    "#1F4E79",  # 深蓝
    "#2F6B9A",  # 中深蓝
    "#4A89B2",  # 中蓝
    "#78A9CC",  # 浅蓝
]

THEME = {
    "figure_bg": "#FFFFFF",
    "axes_bg": "#FFFFFF",
    "stripe_bg": "#FAFBFC",
    "vline_color": "#E3E7EC",
    "hline_color": "#EAEEF2",
    "group_line_color": "#C3CCD6",
    "well_zone_color": "#F4F7FA",
    "bar_edge": "#2F3E4D",
    "label_bbox": "#FFFFFF",
    "title_color": "#1F2F3F",
    "well_label_color": "#35516D",
    "axis_text_color": "#4A5D70",
    "spine_color": "#D2DAE3",
    "leader_line_color": "#8DA0B3",
    "footer_color": "#5D6F82",
}

TYPOGRAPHY = {
    "title_main": 18,
    "title_sub": 12,
    "axis_label": 11,
    "tick_label": 10,
    "well_label": 13,
    "bar_label": 8,
    "bar_label_large": 9,
    "footer": 9,
}

# -----------------------------
# 1️⃣ 找到桌面 Excel 文件
# -----------------------------
desktop_path = Path("/Users/zhouqingquan/Desktop")
files = list(desktop_path.glob("*施工计划表*.xlsx"))
if not files:
    raise FileNotFoundError("桌面上没有找到包含“施工计划表”的Excel文件")
elif len(files) > 1:
    raise FileExistsError("桌面上找到多个包含“施工计划表”的Excel文件，请确保唯一")
excel_file = files[0]
print(f"找到文件：{excel_file}")

# 读取标题（A1 合并单元格）
wb = load_workbook(excel_file, read_only=True)
ws = wb.active
chart_title = ws["A1"].value
print(f"甘特图标题：{chart_title}")

# -----------------------------
# 2️⃣ 读取数据（工程级时间：保留 datetime）
# -----------------------------
df = pd.read_excel(excel_file, header=1)
df = df[["业务类型", "井号", "施工队伍", "施工工序", "开始日期", "结束日期", "时长（天）"]]

df["业务类型"] = df["业务类型"].ffill()
df["井号"] = df["井号"].ffill()
df["施工队伍"] = df["施工队伍"].ffill()

df["开始日期"] = pd.to_datetime(df["开始日期"])
df["结束日期"] = pd.to_datetime(df["结束日期"])

df = df.dropna(subset=["开始日期", "结束日期", "施工工序", "井号", "业务类型"])

bad = df[df["结束日期"] < df["开始日期"]]
if not bad.empty:
    raise ValueError(
        "发现结束日期早于开始日期，请检查Excel：\n"
        + bad[["业务类型", "井号", "施工工序", "开始日期", "结束日期"]].to_string(index=False)
    )

df = df.reset_index(drop=True)

# -----------------------------
# 3️⃣ 工序颜色函数
# -----------------------------
def get_process_color(process_name: str) -> str:
    """
    按工序名称稳定映射到调色盘，不再按“准备/施工/收尾”分类。
    这样视觉更统一，也能保证同名工序跨页面颜色一致。
    """
    key = (process_name or "").strip() or "default"
    digest = hashlib.md5(key.encode("utf-8")).hexdigest()
    color_idx = int(digest[:8], 16) % len(PROCESS_COLOR_PALETTE)
    return PROCESS_COLOR_PALETTE[color_idx]

# -----------------------------
# 4️⃣ 工程级时间引擎
# -----------------------------
def to_engineering_interval(start_dt: pd.Timestamp, end_dt: pd.Timestamp):
    """闭区间（含结束日）→ 右开区间（结束+1天00:00不含）"""
    start = pd.to_datetime(start_dt).normalize()
    end_exclusive = pd.to_datetime(end_dt).normalize() + pd.Timedelta(days=1)
    if end_exclusive <= start:
        raise ValueError(f"日期区间异常：start={start} end={end_dt}")
    return start, end_exclusive

def inclusive_days(start_dt: pd.Timestamp, end_dt: pd.Timestamp) -> int:
    s = pd.to_datetime(start_dt).normalize()
    e = pd.to_datetime(end_dt).normalize()
    return (e - s).days + 1

def fmt_mmdd_dot(ts: pd.Timestamp) -> str:
    return pd.to_datetime(ts).strftime("%m.%d")

# -----------------------------
# ✅ 5️⃣ Excel 单元格二维网格（底纹 + 竖线 + 横线 + 粗分组线）+ 段中心日期标签
# -----------------------------
def set_excel_2d_grid(ax, min_date: pd.Timestamp, max_date: pd.Timestamp, date_span: int, y_total: int, group_separators=None):
    """
    - 交替底纹（像Excel单元格底色）
    - 竖向边框（实线）
    - 横向行边框（实线，形成二维表格）
    - 井与井之间分组边框（更粗横线）
    - 日期标签放在每段中心（底部x轴）
    返回：用于顶部双轴的 locator/formatter 信息
    """
    if group_separators is None:
        group_separators = []

    # 1) 决定“每段长度”（沿用你原规则）
    if date_span <= 40:
        step_days = 1
        fmt = "%m-%d"
    elif date_span <= 60:
        step_days = 2
        fmt = "%m-%d"
    elif date_span <= 200:
        step_days = 5
        fmt = "%m-%d"
    elif date_span <= 365:
        step_days = 7
        fmt = "%m-%d"
    else:
        step_days = 14
        fmt = "%m-%d"

    # 2) 生成段边界
    edges = pd.date_range(start=min_date, end=max_date, freq=f"{step_days}D")
    if len(edges) == 0 or edges[-1] < max_date:
        edges = edges.append(pd.DatetimeIndex([max_date]))
    centers = edges[:-1] + pd.Timedelta(days=step_days / 2)

    edges_num = mdates.date2num(edges.to_pydatetime())
    centers_num = mdates.date2num(centers.to_pydatetime())

    # 3) X轴刻度：major=中心标签，minor=边界
    ax.xaxis.set_major_locator(FixedLocator(centers_num))
    ax.xaxis.set_major_formatter(mdates.DateFormatter(fmt))
    ax.xaxis.set_minor_locator(FixedLocator(edges_num))

    # 4) 画底纹（交替）
    for i in range(len(edges) - 1):
        if i % 2 == 0:
            left = mdates.date2num(edges[i].to_pydatetime())
            right = mdates.date2num(edges[i + 1].to_pydatetime())
            ax.axvspan(left, right, facecolor=THEME["stripe_bg"], alpha=1.0, zorder=0)

    # 5) 画竖线边框（实线）
    y_min = -1
    y_max = max(2, y_total)
    for x in edges_num:
        ax.vlines(x, ymin=y_min, ymax=y_max, colors=THEME["vline_color"], linewidth=0.8, zorder=1)

    # 6) 画横线边框（实线）——二维表格
    x_left = mdates.date2num(min_date.to_pydatetime())
    x_right = mdates.date2num(max_date.to_pydatetime())
    y_lines = np.arange(-0.5, y_total + 0.5, 1.0)
    ax.hlines(y_lines, xmin=x_left, xmax=x_right, colors=THEME["hline_color"], linewidth=0.6, zorder=1)

    # 7) 分组粗线
    for y_sep in group_separators:
        ax.hlines(y_sep, xmin=x_left, xmax=x_right, colors=THEME["group_line_color"], linewidth=1.8, zorder=2)

    # 8) 显示范围
    ax.set_xlim(x_left, x_right)

    return {
        "centers_num": centers_num,
        "fmt": fmt,
        "x_left": x_left,
        "x_right": x_right
    }

def apply_top_axis(ax, tick_info):
    """✅ 顶部双轴：显示同样的日期标签"""
    ax_top = ax.twiny()
    ax_top.set_xlim(tick_info["x_left"], tick_info["x_right"])
    ax_top.xaxis.set_major_locator(FixedLocator(tick_info["centers_num"]))
    ax_top.xaxis.set_major_formatter(mdates.DateFormatter(tick_info["fmt"]))
    ax_top.tick_params(
        axis="x",
        labelrotation=45,
        colors=THEME["axis_text_color"],
        labelsize=TYPOGRAPHY["tick_label"],
    )
    ax_top.grid(False)
    ax_top.spines["top"].set_color(THEME["spine_color"])
    ax_top.spines["top"].set_linewidth(0.8)
    ax_top.spines["bottom"].set_visible(False)
    ax_top.spines["left"].set_visible(False)
    ax_top.spines["right"].set_visible(False)
    return ax_top

# -----------------------------
# ✅ 6️⃣ 井号分区淡红底块
# -----------------------------
def draw_well_zones(ax, well_spans):
    """
    well_spans: [(y_start, y_end), ...]  y_end 为“最后一道工序的下一行”(exclusive)
    画法：覆盖 y_start-0.5 ~ y_end-0.5（正好包住这些行）
    """
    for (ys, ye) in well_spans:
        ax.axhspan(
            ys - 0.5,
            ye - 0.5,
            xmin=0, xmax=1,
            facecolor=THEME["well_zone_color"],  # 井区浅色底块
            alpha=0.55,
            zorder=0.4            # 在灰底纹之上，在网格线/条形之下
        )

# -----------------------------
# 7️⃣ 多施工队伍循环绘图，输出同一 PDF
# -----------------------------
队伍列表 = df["施工队伍"].dropna().unique()
today_str = datetime.now().strftime("%Y-%m-%d")
pdf_path = excel_file.with_name(f"{excel_file.stem}_{today_str}.pdf")
main_title = str(chart_title).strip() if chart_title else "施工进度图"
page_total = len(队伍列表)

with PdfPages(pdf_path) as pdf:
    metadata = pdf.infodict()
    metadata["Title"] = f"{main_title} 甘特图"
    metadata["Author"] = "Gantt Generator"
    metadata["Subject"] = "工程施工进度计划"
    metadata["Creator"] = "gantt_chart/gantt_generator.py"
    metadata["CreationDate"] = datetime.now()
    metadata["ModDate"] = datetime.now()

    for page_idx, 队伍 in enumerate(队伍列表, start=1):
        df_队伍 = df[df["施工队伍"] == 队伍].copy()

        min_date = pd.to_datetime(df_队伍["开始日期"]).min().normalize()
        max_date = pd.to_datetime(df_队伍["结束日期"]).max().normalize() + pd.Timedelta(days=1)
        date_span = (max_date - min_date).days

        fig, ax = plt.subplots(
            figsize=(11.69, 8.27),  # A4 横版固定版式
            facecolor=THEME["figure_bg"]
        )
        ax.set_facecolor(THEME["axes_bg"])

        def format_text_by_days(text, days):
            # 工序名+日期更长，换行适当放宽
            if days >= 6:
                return text, TYPOGRAPHY["bar_label_large"]
            elif days > 1:
                return "\n".join(textwrap.wrap(text, width=14)), TYPOGRAPHY["bar_label"]
            else:
                return "\n".join(textwrap.wrap(text, width=14)), TYPOGRAPHY["bar_label"]

        y_base = 0
        y_ticks = []
        y_labels = []

        group_separators = []   # 分组粗线位置
        well_spans = []         # ✅ 井号分区淡红底块范围：[(y_start, y_end), ...]

        # ---- 先跑数据，记录每口井的行范围 + 画条形 ----
        for 井号, df_井 in df_队伍.groupby("井号", sort=False):
            y_start = y_base        # 该井第一行工序
            井开始_y = y_base

            for _, row in df_井.iterrows():
                proc = str(row["施工工序"])
                color = get_process_color(proc)

                # 工程区间（绘图用）
                start_ts, end_excl_ts = to_engineering_interval(row["开始日期"], row["结束日期"])
                left_num = mdates.date2num(start_ts.to_pydatetime())
                right_num = mdates.date2num(end_excl_ts.to_pydatetime())
                width_days = right_num - left_num

                # 标签：工序 + 日期字段（例：修井车施工准备06.03-06.05）
                label = f"{proc} | {fmt_mmdd_dot(row['开始日期'])}-{fmt_mmdd_dot(row['结束日期'])}"

                # 条形
                ax.barh(
                    y=y_base,
                    width=width_days,
                    left=left_num,
                    height=0.5,
                    color=color,
                    edgecolor=THEME["bar_edge"],
                    linewidth=0.8,
                    zorder=3
                )

                # 文字
                days_inc = inclusive_days(row["开始日期"], row["结束日期"])
                text_str, font_size = format_text_by_days(label, days_inc)

                center_x = left_num + width_days / 2
                right_x = left_num + width_days + 0.2

                if width_days < 1.2:
                    label_x = right_x + 0.2
                    ax.annotate(
                        label,
                        xy=(right_num, y_base),
                        xytext=(label_x, y_base),
                        textcoords="data",
                        ha="left",
                        va="center",
                        fontsize=TYPOGRAPHY["bar_label"],
                        color=THEME["axis_text_color"],
                        arrowprops=dict(
                            arrowstyle="-",
                            color=THEME["leader_line_color"],
                            lw=0.8,
                            shrinkA=0,
                            shrinkB=0,
                        ),
                        zorder=4,
                    )
                else:
                    ax.text(
                        center_x, y_base, text_str,
                        ha="center", va="center",
                        fontsize=font_size, color="#1F2D3D",
                        bbox=dict(facecolor=THEME["label_bbox"], alpha=0.65, edgecolor="none"),
                        zorder=4
                    )

                y_ticks.append(y_base)
                y_labels.append(proc)
                y_base += 1

            y_end = y_base           # ✅ 该井工序结束的下一行（exclusive）
            well_spans.append((y_start, y_end))

            # ✅ 分组粗线：最后一道工序下边界
            group_separators.append(y_end - 0.5)

            # 井与井之间留白一行
            y_base += 1

            # 井号标注（左侧）
            ax.text(
                mdates.date2num(min_date.to_pydatetime()) - 0.5,
                井开始_y - 0.5,
                str(井号),
                ha="right",
                va="bottom",
                fontsize=TYPOGRAPHY["well_label"],
                fontweight="bold",
                color=THEME["well_label_color"],
                zorder=6
            )

        # ---- 画底层：先单元格竖向底纹，再井号淡红分区，再网格线/边框 ----
        tick_info = set_excel_2d_grid(
            ax, min_date, max_date, date_span,
            y_total=y_base,
            group_separators=group_separators
        )

        # ✅ 井号淡红底块（放在灰底纹之上、网格线/条形之下）
        draw_well_zones(ax, well_spans)

        # ✅ 顶部同样日期标签（双轴）
        apply_top_axis(ax, tick_info)

        # 纵轴设置
        ax.set_yticks(y_ticks)
        ax.set_yticklabels(y_labels)
        ax.invert_yaxis()
        ax.tick_params(axis="x", colors=THEME["axis_text_color"], labelsize=TYPOGRAPHY["tick_label"])
        ax.tick_params(axis="y", colors=THEME["axis_text_color"], labelsize=TYPOGRAPHY["tick_label"])
        for spine in ax.spines.values():
            spine.set_color(THEME["spine_color"])
            spine.set_linewidth(0.9)

        # 其他外观
        fig.autofmt_xdate(rotation=45)
        ax.set_xlabel("日期", fontsize=TYPOGRAPHY["axis_label"], color=THEME["axis_text_color"])

        # 标题层级：主标题（工程名）+ 副标题（施工队伍）
        subtitle = f"施工队伍：{队伍}"
        fig.suptitle(
            main_title,
            fontsize=TYPOGRAPHY["title_main"],
            fontweight="bold",
            color=THEME["title_color"],
            y=0.98,
        )
        ax.set_title(
            subtitle,
            fontsize=TYPOGRAPHY["title_sub"],
            fontweight="semibold",
            color=THEME["axis_text_color"],
            pad=10,
        )

        # 页脚：项目名 + 页码
        fig.text(
            0.01, 0.015,
            f"项目：{main_title}",
            ha="left", va="bottom",
            fontsize=TYPOGRAPHY["footer"],
            color=THEME["footer_color"],
        )
        fig.text(
            0.99, 0.015,
            f"第 {page_idx}/{page_total} 页",
            ha="right", va="bottom",
            fontsize=TYPOGRAPHY["footer"],
            color=THEME["footer_color"],
        )

        fig.tight_layout(rect=(0.0, 0.04, 1.0, 0.92))
        fig.subplots_adjust(left=0.25)

        pdf.savefig(fig)
        plt.close(fig)

print(f"✅ 已输出：{pdf_path}")
