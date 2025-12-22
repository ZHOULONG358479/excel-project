from openpyxl import load_workbook
from datetime import datetime
import os
import glob
import re
import sys

# 1. 桌面路径 & 文件前缀
desktop_path = "/Users/zhouqingquan/Desktop"
file_prefix = "第二作业项目部在用承包商员工数统计表"

# 2. 查找 Excel 文件
files = glob.glob(os.path.join(desktop_path, f"{file_prefix}*.xlsx"))
if not files:
    print("❌ 未找到符合条件的 Excel 文件")
    sys.exit(1)

# 3. 取第一个匹配文件
old_path = files[0]
dir_name = os.path.dirname(old_path)
old_name = os.path.basename(old_path)

# 4. 今天日期（用于文件名）
today_cn = datetime.now().strftime("%Y年%m月%d日")

# 5. 重命名文件（替换末尾 8 位数字日期）
new_name = re.sub(r"\d{8}(?=\.xlsx$)", today_cn, old_name)

# 若文件名中本来没有日期
if new_name == old_name:
    new_name = f"{file_prefix}_{today_cn}.xlsx"

new_path = os.path.join(dir_name, new_name)

# 6. 执行重命名
if old_path != new_path:
    os.rename(old_path, new_path)
    print(f"✔ 文件已重命名为：{new_name}")
else:
    print("ℹ 文件名无需修改")

# 7. 打开【重命名后的】Excel
wb = load_workbook(new_path)

# 8. 复制最后一个工作表
last_sheet = wb.worksheets[-1]
new_sheet = wb.copy_worksheet(last_sheet)

# 9. 用 YYYYMMDD 命名新工作表
today_sheet = datetime.now().strftime("%Y%m%d")
if today_sheet in wb.sheetnames:
    del wb[today_sheet]

new_sheet.title = today_sheet
print(f"✔ 已复制最后一个工作表，并命名为：{today_sheet}")

# 10. 修改表头日期
ws = wb[today_sheet]
today_dot = datetime.now().strftime("%Y.%m.%d")

title_cell = ws["A1"]
old_title = title_cell.value

if old_title:
    new_title = re.sub(r"[（(].*?[）)]", f"（{today_dot}）", old_title)
    title_cell.value = new_title
    print("✔ 表头日期已更新")
else:
    print("⚠ A1 单元格为空，未修改表头")

# 11. 保存
wb.save(new_path)
print("🎉 Excel 处理完成")


